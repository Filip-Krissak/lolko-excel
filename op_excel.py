import requests
from bs4 import BeautifulSoup
import openpyxl

# URL of the website to be scraped
url = 'https://www.op.gg/multisearch/euw?summoners=samyaza33'

# Send a GET request to the website and get its HTML content
response = requests.get(url)

# Create a BeautifulSoup object from the HTML content
soup = BeautifulSoup(response.content, 'html.parser')

# Find all the game images in the Recent Games section
game_images = soup.find_all('div', class_='recent-game-image')

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Get the active worksheet
worksheet = workbook.active

# Set the column headings
worksheet['A1'] = 'Match'
worksheet['B1'] = 'Result'
worksheet['C1'] = 'Champion'
worksheet['D1'] = 'KDA'
worksheet['E1'] = 'CS'
worksheet['F1'] = 'Gold'

# Loop through the game images and extract the required information
for i, game_image in enumerate(game_images):
    # Stop after extracting information from the last 5 games
    if i >= 5:
        break

    # Get the champion played by the player
    champion_image = game_image.find('img')
    champion = champion_image['title']

    # Get the KDA, CS, and gold of the player
    game_info_div = game_image.find_next('div', class_='GameItemWrap')
    game_info_span = game_info_div.find_all('span', class_='GameResult')
    result = game_info_span[0].text.strip()

    kda_match = re.search(r'(\d+)/(\d+)/(\d+)', game_info_div.text)
    kda = f'{kda_match.group(1)}/{kda_match.group(2)}/{kda_match.group(3)}'

    cs_match = re.search(r'(\d+) \(\d+\)', game_info_div.text)
    cs = cs_match.group(1)

    gold_match = re.search(r'([0-9,]+)', game_info_div.text)
    gold = gold_match.group(1)

    # Write the information to the worksheet
    worksheet.cell(row=i+2, column=1, value=f'Game {i+1}')
    worksheet.cell(row=i+2, column=2, value=result)
    worksheet.cell(row=i+2, column=3, value=champion)
    worksheet.cell(row=i+2, column=4, value=kda)
    worksheet.cell(row=i+2, column=5, value=cs)
    worksheet.cell(row=i+2, column=6, value=gold)

# Save the workbook to a file
workbook.save('recent_games.xlsx')
